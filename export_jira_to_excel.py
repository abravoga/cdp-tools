#!/usr/bin/env python3
"""
Exportar Jiras asignados a Excel con desplegables de estado y prioridad
"""

from jira_integration import JiraClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime


def get_all_jira_statuses(client):
    """Obtener todos los posibles estados de Jira"""
    # Estados comunes de Jira
    # Podrías también obtenerlos dinámicamente de un issue específico
    return [
        "Open",
        "In Progress",
        "Ready to DEV",
        "In Development",
        "Code Review",
        "Ready for Testing",
        "Testing",
        "Blocked",
        "On Hold",
        "Done",
        "Closed",
        "Cancelled",
        "Resolved"
    ]


def create_jira_excel(output_file="mis_jiras.xlsx"):
    """Crear Excel con Jiras asignados"""

    print("="*80)
    print("EXPORTAR JIRAS A EXCEL")
    print("="*80)

    # Conectar a Jira
    print("\nConectando a Jira...")
    client = JiraClient()

    # Buscar issues asignados
    print("Buscando issues asignados...")
    jql = "assignee = currentUser() AND resolution = Unresolved ORDER BY priority DESC, created DESC"
    results = client.search_issues(jql, max_results=100)

    if not results or 'issues' not in results:
        print("No se encontraron issues")
        return

    issues = results.get('issues', [])
    print(f"Total de issues encontrados: {len(issues)}")

    # Crear workbook
    print("\nCreando Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Mis Jiras"

    # Definir columnas
    headers = [
        "Clave Jira",
        "Resumen",
        "Descripción",
        "Estado Actual",
        "Nuevo Estado",
        "Nueva Prioridad",
        "Reportado por",
        "Enlace"
    ]

    # Estilo de encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Configurar anchos de columna
    ws.column_dimensions['A'].width = 12  # Clave Jira
    ws.column_dimensions['B'].width = 50  # Resumen
    ws.column_dimensions['C'].width = 60  # Descripción
    ws.column_dimensions['D'].width = 18  # Estado Actual
    ws.column_dimensions['E'].width = 18  # Nuevo Estado
    ws.column_dimensions['F'].width = 15  # Nueva Prioridad
    ws.column_dimensions['G'].width = 20  # Reportado por
    ws.column_dimensions['H'].width = 50  # Enlace

    # Lista de estados posibles
    statuses = get_all_jira_statuses(client)
    statuses_str = ",".join(statuses)

    # Lista de prioridades
    priorities = ["P1", "P2", "P3", "P4"]
    priorities_str = ",".join(priorities)

    # Crear validaciones de datos
    status_validation = DataValidation(type="list", formula1=f'"{statuses_str}"', allow_blank=True)
    status_validation.error = 'Selecciona un estado válido'
    status_validation.errorTitle = 'Estado inválido'

    priority_validation = DataValidation(type="list", formula1=f'"{priorities_str}"', allow_blank=True)
    priority_validation.error = 'Selecciona una prioridad válida (P1, P2, P3, P4)'
    priority_validation.errorTitle = 'Prioridad inválida'

    ws.add_data_validation(status_validation)
    ws.add_data_validation(priority_validation)

    # Escribir datos de issues
    print("Procesando issues...")
    for idx, issue in enumerate(issues, 2):  # Empezar en fila 2
        key = issue.get('key', '')
        fields = issue.get('fields', {})

        # Extraer información
        project = fields.get('project', {}).get('key', '')
        issue_type = fields.get('issuetype', {}).get('name', '')
        summary = fields.get('summary', '')

        # Descripción - extraer texto plano de ADF y crear resumen
        description = ''
        desc_content = fields.get('description', {})
        if desc_content and isinstance(desc_content, dict):
            # Intentar extraer texto del formato ADF
            if 'content' in desc_content:
                desc_parts = []
                for content_block in desc_content.get('content', []):
                    if content_block.get('type') == 'paragraph':
                        for text_item in content_block.get('content', []):
                            if text_item.get('type') == 'text':
                                desc_parts.append(text_item.get('text', ''))
                full_description = ' '.join(desc_parts)

                # Crear resumen breve (primera frase o primeros 200 caracteres)
                if full_description:
                    # Buscar el primer punto para obtener la primera frase
                    first_sentence_end = full_description.find('. ')
                    if first_sentence_end > 0 and first_sentence_end < 250:
                        description = full_description[:first_sentence_end + 1]
                    else:
                        # Si no hay punto o está muy lejos, tomar primeros 200 caracteres
                        description = full_description[:200]
                        if len(full_description) > 200:
                            description += '...'
                else:
                    description = summary  # Si no hay descripción, usar el summary

        status = fields.get('status', {}).get('name', '')
        priority_obj = fields.get('priority', {})
        priority = priority_obj.get('name', 'N/A') if priority_obj else 'N/A'
        assignee = fields.get('assignee', {}).get('displayName', 'Unassigned') if fields.get('assignee') else 'Unassigned'
        reporter = fields.get('reporter', {}).get('displayName', 'N/A') if fields.get('reporter') else 'N/A'
        created = fields.get('created', '')[:10]
        updated = fields.get('updated', '')[:10]
        url = f"https://masorange.atlassian.net/browse/{key}"

        # Escribir fila
        row_data = [
            key,
            summary,
            description,  # Ya viene resumida del procesamiento anterior
            status,
            "",  # Nuevo Estado (vacío para que usuario seleccione)
            "",  # Nueva Prioridad (vacío para que usuario seleccione)
            reporter,
            url
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Color alternado para filas
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Hacer enlaces clicables
            if col == 8:  # Columna Enlace
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")

        # Aplicar validación a las celdas de Nuevo Estado y Nueva Prioridad
        status_validation.add(f'E{idx}')  # Columna E: Nuevo Estado
        priority_validation.add(f'F{idx}')  # Columna F: Nueva Prioridad

    # Congelar primera fila
    ws.freeze_panes = 'A2'

    # Guardar archivo
    print(f"\nGuardando archivo: {output_file}")
    wb.save(output_file)

    print("="*80)
    print("EXCEL CREADO EXITOSAMENTE")
    print("="*80)
    print(f"\nArchivo: {output_file}")
    print(f"Issues exportados: {len(issues)}")
    print(f"Columnas: {len(headers)}")
    print("\nCaracterísticas:")
    print("  - Desplegable de estados en columna 'Nuevo Estado' (E)")
    print(f"  - Estados disponibles: {len(statuses)}")
    print("  - Desplegable de prioridades (P1-P4) en columna 'Nueva Prioridad' (F)")
    print("  - URLs clicables a cada issue")
    print("  - Formato con colores y bordes")
    print("  - Primera fila congelada")
    print("\nPuedes abrir el archivo y seleccionar nuevos estados y prioridades")
    print("usando los desplegables en las columnas E y F")
    print("="*80)

    return output_file


if __name__ == "__main__":
    create_jira_excel()
